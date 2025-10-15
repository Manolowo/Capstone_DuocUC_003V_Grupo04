import pandas as pd
import os
import glob
from openpyxl import load_workbook

def extraer_hipervinculos(archivo_excel, celda):
    """Extrae la URL de un hipervínculo en una celda específica"""
    try:
        wb = load_workbook(archivo_excel)
        ws = wb.active
        
        # Convertir coordenadas Excel a índices (ej: 'A1' -> (1, 1))
        if isinstance(celda, str) and celda.replace('$', '').isalpha():
            # Es una referencia de celda como 'A1'
            cell_obj = ws[celda]
        else:
            # Es un índice numérico
            fila, columna = celda
            cell_obj = ws.cell(row=fila, column=columna)
        
        if cell_obj.hyperlink:
            return cell_obj.hyperlink.target
        else:
            return str(cell_obj.value) if cell_obj.value else ''
    except Exception as e:
        return str(cell_obj.value) if 'cell_obj' in locals() and cell_obj.value else ''

def combinar_boletas():
    # Definir la ruta base
    base_path = "C:/Users/diego/Desktop/Nueva carpeta/boletas"
    
    # Encontrar todos los archivos Excel
    patron = os.path.join(base_path, "**", "ventas_por_categorias_*.xlsx")
    archivos = glob.glob(patron, recursive=True)
    
    print(f"🔍 Encontrados {len(archivos)} archivos de boletas")
    
    # Lista para almacenar todos los DataFrames
    todos_los_datos = []
    
    # Procesar cada archivo
    for i, archivo in enumerate(archivos, 1):
        try:
            nombre_archivo = os.path.basename(archivo)
            print(f"📂 Procesando ({i}/{len(archivos)}): {nombre_archivo}")
            
            # PRIMERO: Leer normalmente para obtener la estructura
            df_temp = pd.read_excel(archivo, skiprows=4)
            
            # Verificar que el DataFrame no esté vacío
            if df_temp.empty:
                print(f"   ⚠️  Archivo vacío")
                continue
            
            # IDENTIFICAR columna PDF/hipervínculo
            columnas_pdf = [col for col in df_temp.columns if 'pdf' in str(col).lower() or 
                           'enlace' in str(col).lower() or 'url' in str(col).lower() or
                           'hipervinculo' in str(col).lower()]
            
            if columnas_pdf:
                columna_pdf = columnas_pdf[0]
                print(f"   📎 Columna PDF encontrada: '{columna_pdf}'")
                
                # Crear un nuevo DataFrame para almacenar los hipervínculos
                df_con_hipervinculos = df_temp.copy()
                
                # Cargar el archivo con openpyxl para extraer hipervínculos
                wb = load_workbook(archivo, data_only=False)
                ws = wb.active
                
                # Extraer hipervínculos de la columna PDF
                urls_pdf = []
                for idx in range(len(df_temp)):
                    # +6 porque: skiprows=4 + header (1) + índice base 1
                    fila_excel = idx + 6
                    columna_excel = df_temp.columns.get_loc(columna_pdf) + 1
                    
                    try:
                        cell_obj = ws.cell(row=fila_excel, column=columna_excel)
                        if cell_obj.hyperlink:
                            urls_pdf.append(cell_obj.hyperlink.target)
                        else:
                            urls_pdf.append(str(cell_obj.value) if cell_obj.value else '')
                    except Exception as e:
                        urls_pdf.append(str(df_temp.iloc[idx][columna_pdf]) if pd.notna(df_temp.iloc[idx][columna_pdf]) else '')
                
                # Reemplazar la columna PDF con las URLs extraídas
                df_con_hipervinculos[columna_pdf] = urls_pdf
                
                # Agregar a la lista
                todos_los_datos.append(df_con_hipervinculos)
                print(f"   ✅ {len(df_con_hipervinculos)} registros cargados (con hipervínculos)")
                
            else:
                # Si no hay columna PDF, usar el DataFrame normal
                todos_los_datos.append(df_temp)
                print(f"   ✅ {len(df_temp)} registros cargados (sin columna PDF)")
            
        except Exception as e:
            print(f"   ❌ Error: {str(e)}")
    
    # Combinar todos los DataFrames
    if todos_los_datos:
        print("\n🔄 Combinando todos los datos...")
        df_completo = pd.concat(todos_los_datos, ignore_index=True)
        
        # Mostrar información del dataset combinado
        print(f"\n📊 DATASET COMBINADO:")
        print(f"   Total de registros: {len(df_completo)}")
        print(f"   Total de columnas: {len(df_completo.columns)}")
        
        # Mostrar las primeras columnas
        print(f"\n📋 PRIMERAS COLUMNAS:")
        for i, col in enumerate(df_completo.columns[:10], 1):
            print(f"   {i}. {col}")
        
        if len(df_completo.columns) > 10:
            print(f"   ... y {len(df_completo.columns) - 10} columnas más")
        
        # Verificar si hay columnas de PDF
        columnas_pdf_final = [col for col in df_completo.columns if 'pdf' in str(col).lower() or 
                             'enlace' in str(col).lower() or 'url' in str(col).lower()]
        
        if columnas_pdf_final:
            columna_pdf_final = columnas_pdf_final[0]
            print(f"\n🔗 COLUMNA PDF/HIPERVÍNCULO:")
            print(f"   Columna: '{columna_pdf_final}'")
            print(f"   URLs no vacías: {df_completo[columna_pdf_final].notna().sum()}")
            print(f"   Ejemplos:")
            urls_ejemplo = df_completo[columna_pdf_final].dropna().head(3)
            for url in urls_ejemplo:
                print(f"     - {url}")
        
        # Guardar el archivo combinado
        output_file = "todas_las_boletas_combinadas.xlsx"
        df_completo.to_excel(output_file, index=False)
        
        print(f"\n✅ ARCHIVO GUARDADO: {output_file}")
        print(f"📁 Ubicación: {os.path.abspath(output_file)}")
        
        return df_completo
    else:
        print("❌ No se encontraron datos para combinar")
        return None

# Ejecutar la función
if __name__ == "__main__":
    try:
        # Instalar openpyxl si no está instalado
        try:
            import openpyxl
        except ImportError:
            print("📦 Instalando openpyxl...")
            import subprocess
            subprocess.check_call(["pip", "install", "openpyxl"])
            import openpyxl
        
        boletas_combinadas = combinar_boletas()
        
        if boletas_combinadas is not None:
            print(f"\n📝 MUESTRA DE DATOS (primeras 5 filas):")
            print(boletas_combinadas.head())
            
    except Exception as e:
        print(f"❌ Error general: {str(e)}")